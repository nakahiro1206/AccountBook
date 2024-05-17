import sys
import setting as s
import openpyxl
from openpyxl.styles.fonts import Font
from openpyxl.styles import Alignment
class transaction:
    def __init__(self, List,key):
        if(key=='o'):
            self.date=int(List[s.o.date])-20000000
            self.sub=int(List[s.o.id])-int(List[s.o.date])*10000000
            self.id=int(List[s.o.id])
            self.category=None
            self.name=None
            self.receipt=int(List[s.o.receipt] or 0)
            self.payment=int(List[s.o.payment] or 0)
            self.key=List[s.o.key]
            self.orig_name=List[s.o.origname]
            self.balance=None if(List[s.o.balance] is None) else int(List[s.o.balance])
            self.note=None
            self.link=None
        elif(key=='c'):
            try:
                self.sub=int(List[s.c.sub] or 0)
                self.date=int(List[s.c.date] or 0)
                self.id=(self.date+20000000)*10000000+self.sub
            except:
                self.sub=0
                self.date=List[s.c.date]
                self.id=0
            self.category=List[s.c.category]
            self.name=List[s.c.name]
            self.receipt=int(List[s.c.receipt] or 0)
            self.payment=int(List[s.c.payment] or 0)
            self.key=None
            self.orig_name=None
            self.balance=None
            self.note=List[s.c.note]
            self.link=List[s.c.link]

def CLEAR(row,max_col):
    for index, cell in enumerate(row):
        if(index>=max_col):break
        cell.value=None
        cell.hyperlink=None
    return;

def readsheet(worksheet,minrow,maxcol,tr_lists,key):
    for row in worksheet.iter_rows(min_row=minrow,max_col=maxcol):
        if(row[0].value is None):break
        tr_lists.append(transaction([elem.value for elem in row[0:maxcol]],key=key))
        CLEAR(row,maxcol)

def chobowrite(row,tr,account,row_index,is_minou=False):
    # is_minou: Trueは未納に書く時、Falseは普通の帳簿.
    if(is_minou): 
        row[s.c.id].value=None
        row[s.c.date].value=tr.date if(tr.date!=0) else None
        row[s.c.sub].value=None
        row[s.c.balance].value=None
    else: 
        row[s.c.id].value="=row()-1"
        row[s.c.date].value=tr.date
        row[s.c.sub].value=tr.sub
        row[s.c.balance].value='='+s.c.b_chr+str(row_index-1)+'+'+s.c.r_chr+str(row_index)+'-'+s.c.p_chr+str(row_index)
    row[s.c.category].value=tr.category
    row[s.c.account].value=account
    row[s.c.name].value=tr.name
    row[s.c.receipt].value=tr.receipt
    row[s.c.payment].value=tr.payment
    row[s.c.note].value=tr.note
    row[s.c.link].hyperlink=tr.link
    # date: '0_ '
    row[s.c.sub].number_format='000_);[Red]\\(0\\)'
    # v_sub_colは列の番号にしたので-1する.
    row[s.c.v_sub_col-1].number_format='000_);[Red]\\(0\\)'
    row[s.c.receipt].number_format='#,##0_);[Red]\\(#,##0\\)' #収支報告は'#,##0_);[Red](#,##0)'
    row[s.c.payment].number_format='#,##0_);[Red]\\(#,##0\\)'
    row[s.c.balance].number_format='#,##0_);[Red]\\(#,##0\\)'
    # v_balance_colは列の番号にしたので-1する.
    row[s.c.v_balance_col-1].number_format='#,##0_);[Red]\\(#,##0\\)'

def orig_write(row,tr,_0,_1,_2):
    # _0, _1, _2は使わない変数.
    row[s.o.date].value=tr.date+20000000
    row[s.o.id].value=tr.id
    row[s.o.receipt].value=tr.receipt
    row[s.o.payment].value=tr.payment
    row[s.o.key].value=tr.key
    row[s.o.origname].value=tr.orig_name
    row[s.o.balance].value=tr.balance
    row[s.o.id].number_format='0_ '
    return;

def writesheet(Iter,worksheet,minrow,maxrow,maxcol,writefunc,font,account=None,is_minou=False):
    # worksheet.iter_rowは行ごとの読み取りを行う.
    # 行の範囲はmin_row ~ max_row, 列はmin_col(1 by default) ~ max_col.
    for index,row in enumerate(worksheet.iter_rows(min_row=minrow,max_row=maxrow,max_col=maxcol)):
        try:tr=next(Iter)
        except:break
        # index+minrowで現在の何行目かを示す.
        writefunc(row,tr,account,index+minrow,is_minou)
        # フォント, 左揃え.
        for cell in row:
            cell.font=font
            cell.alignment=Alignment(horizontal="left",vertical="center",wrapText=True)

def lookup(minou_tr_dict,tr,chobo_tr_lists):
    if(tr.name not in minou_tr_dict):return False
    # 帳簿の名前からdatalist作成.
    else: datalist=minou_tr_dict[tr.name]
    #datalistの組み合わせを考える。愚直に2**lengthを全て調べる.
    money=tr.receipt-tr.payment
    length=len(datalist)
    index_list=[i for i in range(1,1<<length)]
    index_list.sort(key=lambda x: bin(x).count("1"))
    for i in index_list:
        tmp_money=0
        for j in range(length):
            if((1<<j)&i):
                tmp_money+=datalist[j].receipt-datalist[j].payment
        if(tmp_money==money):
            print("->帳簿に書き込みます")
            new_datalist=[]
            for j in reversed(range(length)):
                if((1<<j)&i):
                    datalist[j].date=tr.date
                    datalist[j].sub=tr.sub
                    datalist[j].id=tr.id
                    print(
                        "日付: "+str(datalist[j].date)+", 品目: "+str(datalist[j].category)+
                        ", 名前: "+str(datalist[j].name)+", 受取: "+str(datalist[j].receipt)+
                        ", 支払: "+str(datalist[j].payment)+", 備考: "+str(datalist[j].note))
                    chobo_tr_lists.append(datalist[j])
                else: new_datalist.append(datalist[j])
            minou_tr_dict[tr.name]=new_datalist
            return True
    return False

def matchname(tr, dict):
    if(tr.orig_name is None):
        tr.name="ゆうちょ銀行"
        print("変更後の名前: "+str(tr.name))
        if(tr.key=="料\u3000金"):
            tr.category="振込手数料"
            tr.note='="ID: "&ROW()-2'
            return "Yucho"
        elif(tr.key=="受取利子"):
            tr.category="受取利子"
            return "UketoriRishi"
        elif(tr.key=='利子'):return "Rishi"
        elif(tr.key=='税金'):return "Zeikin"
        elif(tr.key=="硬貨料金"):
            tr.category="硬貨料金"
            tr.note='硬貨預け入れの際の手数料'
            return "Yucho"
        else: return "Found";
    if(tr.orig_name in dict):
        tr.name=dict[tr.orig_name]
        print("変更後の名前: "+str(tr.name))
        return "Found"
    return "NotFound";

def MergeCells(chobo):
    date_start_row=s.c.min_row;date_tmp=-1
    sub_start_row=s.c.min_row;sub_tmp=-1
    for index,row in enumerate(chobo.iter_rows(min_row=s.c.min_row,max_row=s.c.max_row,max_col=s.c.max_col)):
        date_new=int(row[s.c.date].value or -1)
        sub_new=int(row[s.c.sub].value or -1)
        if(date_tmp==-1):
            date_tmp=int(date_new)
            sub_tmp=int(sub_new)
            continue;
        if(date_tmp!=date_new):
            # 結合と書き込み.
            chobo.merge_cells(start_row=date_start_row,end_row=index+s.c.min_row-1,start_column=s.c.v_date_col,end_column=s.c.v_date_col)
            chobo.cell(row=date_start_row,column=s.c.v_date_col,value=date_tmp)
            date_start_row=index+s.c.min_row
            date_tmp=date_new
            # subも更新する.
            chobo.merge_cells(start_row=sub_start_row,end_row=index+s.c.min_row-1,start_column=s.c.v_sub_col,end_column=s.c.v_sub_col)
            chobo.merge_cells(start_row=sub_start_row,end_row=index+s.c.min_row-1,start_column=s.c.v_balance_col,end_column=s.c.v_balance_col)
            chobo.merge_cells(start_row=sub_start_row,end_row=index+s.c.min_row-1,start_column=s.c.v_name_col,end_column=s.c.v_name_col)
            chobo.cell(row=sub_start_row,column=s.c.v_sub_col,value=sub_tmp)
            chobo.cell(row=sub_start_row,column=s.c.v_balance_col,value="="+s.c.b_chr+str(index+s.c.min_row-1))
            chobo.cell(row=sub_start_row,column=s.c.v_name_col,value="="+s.c.n_chr+str(index+s.c.min_row-1))
            sub_start_row=index+s.c.min_row
            sub_tmp=sub_new
            continue
        if(sub_tmp!=sub_new):
            chobo.merge_cells(start_row=sub_start_row,end_row=index+s.c.min_row-1,start_column=s.c.v_sub_col,end_column=s.c.v_sub_col)
            chobo.merge_cells(start_row=sub_start_row,end_row=index+s.c.min_row-1,start_column=s.c.v_balance_col,end_column=s.c.v_balance_col)
            chobo.merge_cells(start_row=sub_start_row,end_row=index+s.c.min_row-1,start_column=s.c.v_name_col,end_column=s.c.v_name_col)
            chobo.cell(row=sub_start_row,column=s.c.v_sub_col,value=sub_tmp)
            chobo.cell(row=sub_start_row,column=s.c.v_balance_col,value="="+s.c.b_chr+str(index+s.c.min_row-1))
            chobo.cell(row=sub_start_row,column=s.c.v_name_col,value="="+s.c.n_chr+str(index+s.c.min_row-1))
            sub_start_row=index+s.c.min_row
            sub_tmp=sub_new
        if(date_new<0):break

def main(workbook,dict,account=None):
    orig=workbook["入出金明細(原本)"]
    chobo=workbook["帳簿"]
    minou=workbook["未納・未返金"]
    untreated=workbook["未処理(原本)"]

    untreated_tr_lists=[] # 未処理原本をストック.
    untreated_tr_result=[] # 未処理原本に書き込む用.
    unregisterd_name=[] # 未登録の名前.
    minou_tr_dict={} # 未納欄の取引を名前ごとに登録
    chobo_tr_lists=[] # 帳簿に書き込む用.
    orig_done_lists=[] # 原本に書き込む用.
    complete=True
    # 帳簿の結合セルを全て解除.
    for RangeString in list(chobo.merged_cells.ranges):
        chobo.unmerge_cells(range_string=str(RangeString))
    # 原本読み込み
    readsheet(orig,s.o.min_row,s.o.max_col,orig_done_lists,'o')
    # 帳簿読み込み(前年度引き継ぎの行は飛ばす)
    readsheet(chobo,s.c.min_row,s.c.max_col,chobo_tr_lists,'c')
    # 未処理原本読み込み
    readsheet(untreated,s.o.min_row,s.o.max_col,untreated_tr_lists,'o')
    # 逆向きの辞書を作成.
    # dict: A-> C, B->C. reverse_dict: C->[A, B]
    reverse_dict={}
    for t in dict.items():
        key=t[0];value=t[1]
        if(value  not in reverse_dict): reverse_dict[value]=[key]
        else: reverse_dict[value].append(key)
    # 未納欄からリストの辞書を作成
    for row in minou.iter_rows(min_row=s.minou_min_row, max_col=s.c.max_col):
        name=row[s.c.name].value
        if(name is None):break
        tr=transaction([elem.value for elem in row[0:s.c.max_col]],key='c')
        # 品目か備考がWRITE_HEREのままなら保持せずに削除.
        if(tr.category!="WRITE_HERE" and tr.note!="WRITE_HERE"):
            # 変更前の名前で未納欄に書かれていた場合、名前を変更.
            if(name in dict): tr.name=dict[name]
            if(name not in minou_tr_dict):
                minou_tr_dict[name]=[tr]
            else: minou_tr_dict[name].append(tr)
        CLEAR(row,s.c.max_col)
    
    # ここから処理部分.
    # 未処理原本を処理.
    UketoriRishiTr=None # 受取利子をここに保留させる.
    for tr in untreated_tr_lists:
        print("_"*60)
        print("日付: "+str(tr.date)+", 名前: "+str(tr.orig_name)+", 受取: "+str(tr.receipt)+", 支払: "+str(tr.payment))
        # 対応表をもとに名前を見つける
        match_result=matchname(tr,dict)
        # UketoriRishiTrの処理。受取利子のあとに利子、税金が続かないケース.
        if(UketoriRishiTr is not None and match_result!="Rishi" and match_result!="Zeikin"):
            chobo_tr_lists.append(UketoriRishiTr)
            UketoriRishiTr=None
        
        if(match_result=="NotFound"):
            # 名前が登録されていない。未処理欄に書き込む.
            print("->未処理欄に書き込みます")
            tr.name=tr.orig_name
            untreated_tr_result.append(tr)
            unregisterd_name.append(tr.orig_name)
            complete=False
        elif(match_result=="Rishi"):
            # 利子の処理。帳簿には書き込まず原本にのみ書き込む.
            print("->原本に書き込みます")
            orig_done_lists.append(tr)
            UketoriRishiTr.note="利子が"+str(tr.receipt)+"円、"
        elif(match_result=="Zeikin"):
            # 税金の処理。帳簿には書き込まず原本にのみ書き込む.
            print("->原本に書き込みます")
            orig_done_lists.append(tr)
            UketoriRishiTr.note+="税金が"+str(tr.payment)+"円。"
            chobo_tr_lists.append(UketoriRishiTr)
            UketoriRishiTr=None
        elif(match_result=="UketoriRishi"):
            # 受取利子の処理。原本にのみ書き込み、帳簿は一旦保留.
            print("->原本に書き込みます")
            orig_done_lists.append(tr)
            UketoriRishiTr=tr
        elif(match_result=="Yucho"):
            # ゆうちょの手数料や受取利子.
            chobo_tr_lists.append(tr)
            orig_done_lists.append(tr)
        elif(match_result=="Found"):
            # 一般的な処理。未納欄からlookupする
            found=lookup(minou_tr_dict,tr,chobo_tr_lists)
            if(found==False):
                # 該当する取引が見つからなかった場合.
                print("->未処理欄に書き込みます")
                untreated_tr_result.append(tr)
                complete=False
            else:
                orig_done_lists.append(tr)
        else:sys.exit("matchname参照.ゆうちょの特殊な処理の可能性あり")
    if(UketoriRishiTr is not None):
        chobo_tr_lists.append(UketoriRishiTr)
        UketoriRishiTr=None
    
    # 書き込みを行う.
    font=Font(name='Calibri',size=11,bold=False,italic=False,underline='none',strike=False,color='FF000000')
    chobo_tr_lists.sort(key=lambda tr: tr.id)
    orig_done_lists.sort(key=lambda tr: tr.id)
    # 帳簿書き込み.
    writesheet(iter(chobo_tr_lists),chobo,s.c.min_row,s.c.max_row,s.c.max_col,chobowrite,font,account)
    # 原本書き込み.
    writesheet(iter(orig_done_lists),orig,s.o.min_row,s.o.max_row,s.o.max_col,orig_write,font)
    # 未処理原本書き込み.
    writesheet(iter(untreated_tr_result),untreated,s.o.min_row,s.o.max_row,s.o.max_col,orig_write,font)
    # 未納書き込み.
    minou_tr_list=[]
    for tr_list in minou_tr_dict.values():minou_tr_list+=tr_list
    # 実行後に編集しやすいよう未処理原本を未納欄に反映.
    for tr in untreated_tr_result:
        tr.category="WRITE_HERE";tr.note="WRITE_HERE"
        minou_tr_list.append(tr)
    writesheet(iter(minou_tr_list),minou,s.minou_min_row,s.c.max_row,s.c.max_col,chobowrite,font,account,is_minou=True)
    # 未登録の名前をNoneで追加.
    for name in unregisterd_name: dict[name]=None 
    # mergecell.
    MergeCells(chobo)
    print("実行終了(未保存)")
    return complete, dict