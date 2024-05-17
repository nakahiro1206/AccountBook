import openpyxl
import sys

"""
部員の名前はB列の3行目から
C列は総額
E, H, Kは備考
F, I, Lは金額(数式はNG)
最後は過去の集金過不足分で閉じる
"""

filepath="/Users/nakanohiroki/Downloads/230601集金シート.xlsx"
tantou=None
workbook=openpyxl.load_workbook(filepath)
shukin=workbook["集金シート"]
chobo=workbook["帳簿用"]
number=int(input("過去の集金過不足分を除いて、何種類の集金項目がありますか。"))
if(shukin[1][3*number+4].value!="過去の集金過不足分"):
    sys.exit("集金の項目数が合っていません")
categorylist=[]
for i in range(number):
    categorylist.append(shukin[1][3*i+4].value)
print(categorylist,"(+過去の集金過不足分)")
while(1):
    index=input("変更があるならそのindex, なければEnter")
    if(index==""): break
    newcategory=input("新しい品目")
    categorylist[int(index)]=newcategory
    print(categorylist)
#example: B4=[4][1]
write_row=1
write_list=[]
for row in shukin.iter_rows(min_row=3):
    name=row[1].value
    if(name is None): break
    for i in range(number):
        amount=int(row[3*i+5].value or 0)
        if(amount==0): continue
        uketori=amount if(amount>0) else 0
        shiharai=-amount if(amount<0) else 0
        category=categorylist[i]
        bikou=row[3*i+4].value
        #chobo{(id, date, sub,) 内容, 担当, 相手, 受け取り, 支払い, 残高, 備考}
        write_list.append([category,tantou,name,uketori,shiharai,None,bikou])

write_list_iter=iter(write_list)
for row in chobo.iter_rows(min_row=1,max_row=500, max_col=7):
    try: write_list_cell_iter=iter(next(write_list_iter))
    except:
        break
    for cell in row:
        try: 
            cell.value=next(write_list_cell_iter)
        except:
            break
workbook.save(filepath)
