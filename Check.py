import openpyxl
import sys
import setting as s

class ChoboClass:
    def __init__(self,chobo_tuple):
        self.id=(int(chobo_tuple[s.c.date].value or 0)+20000000)*10000000+int(chobo_tuple[s.c.sub].value or 0)
        self.name=chobo_tuple[s.c.name].value
        self.amount=int(chobo_tuple[s.c.receipt].value or 0)-int(chobo_tuple[s.c.payment].value or 0)

class OrigClass:
    def __init__(self, orig_tuple):
        self.id=int(orig_tuple[s.o.id].value or 0)
        self.key=orig_tuple[s.o.key].value
        self.name=orig_tuple[s.o.origname].value
        self.balance=int(orig_tuple[s.o.balance].value or 0)
        self.amount=int(orig_tuple[s.o.receipt].value or 0)-int(orig_tuple[s.o.payment].value or 0)

def APPEND(keisan_writelist,c,o,sub_is_None):
    # keisan[keisan_row][0].number_format='0_ '
    keisan_writelist.append([c.id,
        0 if(sub_is_None) else o.amount,
        o if(sub_is_None) else o.balance,
        None if(sub_is_None) else o.name,
        c.name]);

def PrintTr(msg, c, o, tmp_balance=False):
    balance="省略" if(tmp_balance==False) else str(tmp_balance)
    MSG="State: "+str(msg)
    if(msg=="名前の不一致"): MSG+="\n帳簿とその次の行の情報\nID: "
    else: MSG+="\n帳簿と原本の情報\nID: "
    MSG+=str(c.id)+" "*15+str(o.id)
    MSG+="\n名前: "+str(c.name)+" "*15+str(o.name)
    MSG+="\n取引額: "+str(c.amount)+" "*15+str(o.amount)
    if(msg!="名前の不一致"): MSG+="\n残高: "+str(balance)+" "*15+str(o.balance)
    return MSG

def MakeDict(match,keisan=False):
    dict={}
    # 対応表からdict作成.
    for row in match.iter_rows(max_col=2,max_row=s.match_max_row):
        if(row[0].value is None):break
        if(row[0].value not in dict):
            if(row[1].value is not None):
                dict[row[0].value]=row[1].value
    # 計算シートからdict作成.
    if(keisan!=False):
        for row in keisan.iter_rows(min_row=s.keisan_min_row, max_col=5,max_row=s.keisan_max_row):
            before=row[3].value
            if(before is None):continue
            after=row[4].value
            if(before not in dict):
                dict[before]=after
            elif(after!=dict[before]):
                print(before,"に対して",dict[before],"が既に登録されているので",after,"は登録されません")
    return dict

def WriteMatch(match,dict):
    # 対応表書き込み.
    dict_item_list=sorted(dict.items(),key=lambda x:str(x[1] or ''))
    dict_len=len(dict_item_list)
    dict_item_iter=iter(dict_item_list)
    for index,row in enumerate(match.iter_rows(max_col=2,max_row=s.match_max_row)):
        if(index>=dict_len):
            row[0].value=None
            row[1].value=None
        else:
            item=next(dict_item_iter)
            row[0].value=item[0]
            row[1].value=item[1]

def main(fp=False,wb=False,save=False,complete=False,newdict=False):
    if(wb!=False): workbook=wb
    elif(fp==False):
        sys.exit("引数エラー")
    else:
        print("Excelファイルの読み込み中...")
        workbook=openpyxl.load_workbook(fp)
    print("原本と帳簿の照合を開始します")
    orig=workbook["入出金明細(原本)"]
    chobo=workbook["帳簿"]
    keisan=workbook["計算"]
    match=workbook["対応表"]
    keisan_writelist=[["取引ID","金額","残高","名前(原本)","名前(帳簿)"]]
    keisan_maxcol=len(keisan_writelist[0])
    # iterator.
    orig_iter=orig.iter_rows(min_row=s.o.min_row,max_row=s.o.max_row)
    chobo_iter=chobo.iter_rows(min_row=s.c.min_row,max_row=s.c.max_row)
    # 前年度引き継ぎの金額.
    tmp_balance=int(chobo[s.c.b_chr+str(s.c.min_row-1)].value or 0)
    #原本の取引
    o=OrigClass(next(orig_iter)[0:s.o.max_col])
    #帳簿の取引をまとめる
    c=ChoboClass(next(chobo_iter)[0:s.c.max_col])
    # 帳簿、原本がまっさら(年始など)の場合IterationStopするのでこの処理を追加
    if(c.id==200000000000000 and o.id==0):
        dict=MakeDict(match)
        return workbook, dict
    #ncは次の行
    nc=ChoboClass(next(chobo_iter)[0:s.c.max_col])
    while(1):
        #利子税金を弾く
        if(o.key=="利子"):
            # 利子、税金、次の取引なので2つとばす.
            _=next(orig_iter)
            o=OrigClass(next(orig_iter)[0:s.o.max_col])
        
        tmp_balance+=c.amount
        assert c.id<=nc.id or nc.id==200000000000000, "ID: "+str(c.id)+"昇順にしてください"
        # 金額が相殺され、集金は行っていないが取引は記入したいケース
        if(c.id%10000000==0):
            # idが同じ処理をまとめる.
            while(c.id==nc.id):
                assert c.name==nc.name, PrintTr("名前の不一致", c, nc)
                c.amount+=nc.amount
                tmp_balance+=nc.amount
                nc=ChoboClass(next(chobo_iter)[0:s.c.max_col])
            assert c.amount==0, str(c.id)+"で合計が0にならず"+str(c.amount)+"です"
            APPEND(keisan_writelist, c, tmp_balance, sub_is_None=True)
            c=nc
            nc=ChoboClass(next(chobo_iter)[0:s.c.max_col])
            continue;
        # idが同じ処理をまとめる.
        while(c.id==nc.id):
            assert c.name==nc.name, PrintTr("名前の不一致", c, nc)
            c.amount+=nc.amount
            tmp_balance+=nc.amount
            nc=ChoboClass(next(chobo_iter)[0:s.c.max_col])
        
        # 原本と照合
        assert o.id==c.id, PrintTr("IDの不一致", c, o)
        assert o.amount==c.amount, PrintTr("取引額の不一致", c, o)
        if(complete==True): assert o.balance==tmp_balance, PrintTr("残高の不一致", c, o,tmp_balance)
        # テスト用.
        # print(PrintTr("OK", c, o))
        APPEND(keisan_writelist, c, o, sub_is_None=False)
        c=nc
        nc=ChoboClass(next(chobo_iter)[0:s.c.max_col])
        o=OrigClass(next(orig_iter)[0:s.o.max_col])
        if(o.id==0):print("チェック完了");break
    
    # 計算シートに書き込み
    keisan_writeiter=iter(keisan_writelist)
    for row in keisan.iter_rows(max_col=keisan_maxcol):
        try: elem=next(keisan_writeiter)
        except: break
        row[0].number_format='0_ '
        for i in range(keisan_maxcol):
            row[i].value=elem[i]
    # 対応表作成.
    dict=newdict if(newdict!=False) else MakeDict(match,keisan)
    WriteMatch(match,dict)
    print("実行完了(未保存)")
    if(save):
        if(complete):
            print(fp+"\nをまとめに変更してもよろしいですか?")
            if(not len(input("よければEnter"))):
                orig.sheet_state='hidden'
                keisan.sheet_state='hidden'
                match.sheet_state='hidden'
                workbook["未処理(原本)"].sheet_state='hidden'
                print("まとめを作成しました\nピボットテーブルを更新してください")
            else: 
                orig.sheet_state='visible'
                keisan.sheet_state='visible'
                match.sheet_state='visible'
                workbook["未処理(原本)"].sheet_state='visible'
                print("帳簿作成完了")
        else: 
            orig.sheet_state='visible'
            keisan.sheet_state='visible'
            match.sheet_state='visible'
            workbook["未処理(原本)"].sheet_state='visible'
            print('帳簿の未納、未返金欄を追加してください')
        workbook.save(fp)
        print("保存しました")
        return;
    else: return workbook, dict